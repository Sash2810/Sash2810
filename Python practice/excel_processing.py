import openpyxl as xl
from openpyxl.chart import BarChart, Reference   
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
sheet.cell(1, 1)
print(sheet.max_column)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_value = cell.value * 0.9
    corrected_value_cell = sheet.cell(row, 4)
    corrected_value_cell.value = corrected_value
    
    
values = Reference(sheet, min_row = 2,
          max_row = sheet.max_row,
          min_col = 4, max_col = 4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'm25')
    
wb.save('Transaction.xlsx')